import openpyxl

# Create a new workbook and select the active sheet
wb = openpyxl.Workbook()
ws = wb.active

# Add headers
ws['A1'] = 'Description'

# Sample security threats
threats = [
    "A remote attacker can execute arbitrary code via a specially crafted HTTP request to the web server.",
    "Local users can gain elevated privileges by exploiting a buffer overflow in the system service.",
    "An attacker with physical access can extract sensitive data from unencrypted backup files.",
    "A SQL injection vulnerability in the login form allows unauthorized access to the database.",
    "Cross-site scripting (XSS) vulnerability in the comment section enables attackers to inject malicious scripts."
]

# Add threats to the worksheet
for i, threat in enumerate(threats, start=2):  # Start from row 2 (after header)
    ws[f'A{i}'] = threat

# Save the workbook
wb.save('sample_threats.xlsx')
print("Sample Excel file created successfully!")
