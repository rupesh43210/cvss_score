import os
import json
import logging
import pandas as pd
import openpyxl
from typing import Dict, List, Optional
from datetime import datetime
from werkzeug.utils import secure_filename
from flask import current_app
from cvss.utils.logger import setup_logger
import csv

logger = setup_logger(__name__)

class FileService:
    """Service for handling file operations."""
    
    ALLOWED_EXTENSIONS = {'xlsx', 'xls', 'csv'}
    
    def __init__(self):
        """Initialize the FileService."""
        self._upload_folder = None
        self._output_folder = None

    @property
    def upload_folder(self) -> str:
        """Get the upload folder path."""
        if self._upload_folder is None:
            self._upload_folder = current_app.config['UPLOAD_FOLDER']
            os.makedirs(self._upload_folder, exist_ok=True)
        return self._upload_folder

    @property
    def output_folder(self) -> str:
        """Get the output folder path."""
        if self._output_folder is None:
            self._output_folder = current_app.config['OUTPUT_FOLDER']
            os.makedirs(self._output_folder, exist_ok=True)
        return self._output_folder

    def _allowed_file(self, filename: str) -> bool:
        """Check if the file extension is allowed."""
        return '.' in filename and filename.rsplit('.', 1)[1].lower() in self.ALLOWED_EXTENSIONS

    def _generate_safe_filename(self, original_filename: str, prefix: str = "") -> str:
        """Generate a safe filename with timestamp."""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = secure_filename(original_filename)
        name, ext = os.path.splitext(filename)
        return f"{prefix}{name}_{timestamp}.xlsx"  # Always use .xlsx extension

    def save_uploaded_file(self, file) -> Optional[str]:
        """Save the uploaded file and return its path."""
        try:
            if not file or not file.filename:
                logger.error("No file provided")
                return None

            if not self._allowed_file(file.filename):
                logger.error(f"Invalid file type: {file.filename}")
                return None

            filename = self._generate_safe_filename(file.filename)
            filepath = os.path.join(self.upload_folder, filename)
            
            # Save the file
            file.save(filepath)
            logger.info(f"File saved successfully: {filepath}")
            
            return filename

        except Exception as e:
            logger.error(f"Error saving file: {str(e)}", exc_info=True)
            return None

    def create_excel_output(self, results: List[Dict], base_filepath: str) -> Optional[str]:
        """Create an Excel file with the analysis results."""
        try:
            # Generate safe filename for output
            output_filename = self._generate_safe_filename(os.path.basename(base_filepath), prefix="analysis_")
            output_path = os.path.join(self.output_folder, output_filename)

            # Prepare data for Excel
            data = []
            for result in results:
                metrics = result['metrics']
                data.append({
                    'Description': result['description'],
                    'CVSS Score': result['cvss_score'],
                    'Severity': result['severity'],
                    'Vector String': f"CVSS:3.1/AV:{metrics['AV']}/AC:{metrics['AC']}/PR:{metrics['PR']}/UI:{metrics['UI']}/S:{metrics['S']}/C:{metrics['C']}/I:{metrics['I']}/A:{metrics['A']}",
                    'Justification': metrics.get('justification', ''),
                    'Explanation': metrics.get('explanation', ''),
                    'Attack Vector': self._get_full_metric_name('AV', metrics['AV']),
                    'Attack Complexity': self._get_full_metric_name('AC', metrics['AC']),
                    'Privileges Required': self._get_full_metric_name('PR', metrics['PR']),
                    'User Interaction': self._get_full_metric_name('UI', metrics['UI']),
                    'Scope': self._get_full_metric_name('S', metrics['S']),
                    'Confidentiality': self._get_full_metric_name('C', metrics['C']),
                    'Integrity': self._get_full_metric_name('I', metrics['I']),
                    'Availability': self._get_full_metric_name('A', metrics['A']),
                    'Confidence': f"{metrics.get('confidence', 'N/A')}%"
                })

            # Create DataFrame and write to Excel
            df = pd.DataFrame(data)
            
            # Reorder columns to match display order
            column_order = [
                'Description',
                'CVSS Score',
                'Severity',
                'Vector String',
                'Justification',
                'Explanation',
                'Attack Vector',
                'Attack Complexity',
                'Privileges Required',
                'User Interaction',
                'Scope',
                'Confidentiality',
                'Integrity',
                'Availability',
                'Confidence'
            ]
            df = df[column_order]
            
            # Write to Excel with formatting
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='CVSS Analysis')
                
                # Get the workbook and the worksheet
                workbook = writer.book
                worksheet = writer.sheets['CVSS Analysis']
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column[0].column_letter].width = min(adjusted_width, 50)
                
                # Add text wrapping for description and justification columns
                wrap_text_style = openpyxl.styles.Alignment(wrap_text=True, vertical='top')
                for row in worksheet.iter_rows(min_row=2):  # Skip header row
                    row[0].alignment = wrap_text_style  # Description
                    row[4].alignment = wrap_text_style  # Justification
                    row[5].alignment = wrap_text_style  # Explanation
                
                # Style the header row
                header_style = openpyxl.styles.NamedStyle(name='header')
                header_style.font = openpyxl.styles.Font(bold=True)
                header_style.fill = openpyxl.styles.PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
                header_style.alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')
                
                for cell in worksheet[1]:
                    cell.style = header_style
                
                # Set row height for wrapped text
                for row in worksheet.iter_rows(min_row=2):
                    worksheet.row_dimensions[row[0].row].height = 60  # Adjust this value as needed

            logger.info(f"Excel file created successfully at: {output_path}")
            return output_filename

        except Exception as e:
            logger.error(f"Error creating Excel file: {str(e)}", exc_info=True)
            return None

    def _get_full_metric_name(self, metric: str, value: str) -> str:
        """Convert metric abbreviations to full names."""
        metric_names = {
            'AV': {
                'N': 'Network',
                'A': 'Adjacent',
                'L': 'Local',
                'P': 'Physical'
            },
            'AC': {
                'L': 'Low',
                'H': 'High'
            },
            'PR': {
                'N': 'None',
                'L': 'Low',
                'H': 'High'
            },
            'UI': {
                'N': 'None',
                'R': 'Required'
            },
            'S': {
                'U': 'Unchanged',
                'C': 'Changed'
            },
            'C': {
                'N': 'None',
                'L': 'Low',
                'H': 'High'
            },
            'I': {
                'N': 'None',
                'L': 'Low',
                'H': 'High'
            },
            'A': {
                'N': 'None',
                'L': 'Low',
                'H': 'High'
            }
        }
        return metric_names.get(metric, {}).get(value, value)

    def read_threats(self, filepath: str) -> Optional[List[Dict]]:
        """Read threats from the uploaded file."""
        try:
            # Read the file based on its extension
            _, ext = os.path.splitext(filepath)
            ext = ext.lower()

            if ext == '.csv':
                threats = []
                with open(filepath, 'r', encoding='utf-8') as f:
                    reader = csv.DictReader(f)
                    # Try various possible column names
                    description_variants = ['Description', 'description', 'Threat', 'threat',
                                         'Vulnerability', 'vulnerability', 'Details', 'details',
                                         'Issue', 'issue', 'Finding', 'finding', 'Risk', 'risk']
                    
                    # Find the actual column name from the CSV
                    description_col = None
                    for variant in description_variants:
                        if variant in reader.fieldnames:
                            description_col = variant
                            break
                    
                    if description_col is None:
                        logger.error("No valid description column found in CSV file")
                        return None
                    
                    for row in reader:
                        description = row.get(description_col, '').strip()
                        if description:  # Only add non-empty descriptions
                            threats.append({
                                'description': description
                            })
                return threats if threats else None
            else:
                df = pd.read_excel(filepath, engine='openpyxl')
                
                # List of possible column names for threat description
                description_variants = [
                    'description', 'threat', 'vulnerability', 'details',
                    'issue', 'finding', 'risk', 'summary', 'overview',
                    'desc', 'title', 'name', 'threat description',
                    'vulnerability description', 'security issue'
                ]
                
                # Find the description column (case-insensitive)
                description_col = None
                df.columns = df.columns.str.lower()  # Convert all column names to lowercase
                
                # First try exact matches
                for col in df.columns:
                    if col in description_variants:
                        description_col = col
                        break
                
                # If no exact match, try partial matches
                if description_col is None:
                    for col in df.columns:
                        for variant in description_variants:
                            if variant in col or col in variant:
                                description_col = col
                                break
                        if description_col:
                            break
                
                if description_col is None:
                    # If still no match found, try the first text-like column
                    for col in df.columns:
                        # Check if the column contains mostly text data
                        if df[col].dtype == 'object':
                            sample = df[col].dropna().head()
                            if len(sample) > 0 and all(isinstance(x, str) and len(x.strip()) > 0 for x in sample):
                                description_col = col
                                logger.info(f"Using column '{col}' as description column based on content analysis")
                                break
                
                if description_col is None:
                    logger.error("No suitable description column found in Excel file")
                    return None

                # Extract threats
                threats = []
                for _, row in df.iterrows():
                    description = row[description_col]
                    if pd.notna(description) and str(description).strip():
                        description = str(description).strip()
                        if len(description) > 0:  # Additional check for non-empty strings
                            threats.append({
                                'description': description
                            })

                return threats if threats else None

        except Exception as e:
            logger.error(f"Error reading threats from file: {str(e)}", exc_info=True)
            return None
